"""Generate payment_data.json from a merged _汇总.xlsx file."""
import openpyxl, json, sys
from openpyxl.cell.cell import MergedCell
from datetime import datetime, date


def safe_read(cell):
    if isinstance(cell, MergedCell):
        return None
    return cell.value


def to_num(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    cleaned = s.replace('$', '').replace('£', '').replace('€', '').replace('¥', '').replace(',', '').replace(' ', '')
    try:
        return float(cleaned) if '.' in cleaned else int(cleaned)
    except:
        return s


def fmt_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%m/%d/%Y')
    if isinstance(val, date):
        return val.strftime('%m/%d/%Y')
    s = str(val).strip()
    if not s:
        return None
    # Already a string like "08/11/2026"
    return s


def build(xlsx_path, pull_time):
    wb = openpyxl.load_workbook(xlsx_path)
    sheets = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = []
        for r in range(2, ws.max_row + 1):
            a = safe_read(ws.cell(r, 1))
            b = safe_read(ws.cell(r, 2))
            c = safe_read(ws.cell(r, 3))
            d = safe_read(ws.cell(r, 4))
            e = safe_read(ws.cell(r, 5))
            f = safe_read(ws.cell(r, 6))
            if all(v is None for v in [a, b, c, d, e, f]):
                continue
            rows.append({
                'A': str(a).strip() if a is not None else '',
                'B': str(b).strip() if b is not None else '',
                'C': to_num(c),
                'D': fmt_date(d),
                'E': to_num(e),
                'F': to_num(f),
            })
        sheets[sn] = rows
    return {'pull_time': pull_time, 'sheets': sheets}


if __name__ == '__main__':
    xlsx = sys.argv[1]
    out = sys.argv[2]
    pt = sys.argv[3] if len(sys.argv) > 3 else datetime.now().strftime('%Y-%m-%d %H:%M')
    data = build(xlsx, pt)
    with open(out, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    total = sum(len(rows) for rows in data['sheets'].values())
    print(f"Wrote {out}: {len(data['sheets'])} sheets, {total} rows, pull_time={pt}")
    for sn, rows in data['sheets'].items():
        print(f"  [{sn}] {len(rows)} rows")
